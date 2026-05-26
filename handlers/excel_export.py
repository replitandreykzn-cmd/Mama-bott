"""
Экспорт данных ребёнка в Excel (.xlsx).
"""
import io
from datetime import date, datetime

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
import database as db
from handlers.child import age_str


def generate_child_xlsx(child, growth_records, vaccinations,
                        illnesses=None, medications=None, allergies=None, contraindications=None) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, GradientFill
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl не установлен. Добавьте openpyxl в requirements.txt")

    wb = openpyxl.Workbook()

    # ── Цвета ─────────────────────────────────────────────────────────────
    PURPLE      = "FF6366F1"   # accent
    PURPLE_LIGHT= "FFEFE6FF"
    GREEN       = "FF22C55E"
    GREEN_LIGHT = "FFF0FDF4"
    AMBER       = "FFF59E0B"
    AMBER_LIGHT = "FFFFFBEB"
    PINK        = "FFEC4899"
    PINK_LIGHT  = "FFFDF2F8"
    GRAY_DARK   = "FF111827"
    GRAY_MID    = "FF6B7280"
    GRAY_LIGHT  = "FFF9FAFB"
    WHITE       = "FFFFFFFF"
    RED         = "FFEF4444"

    def hdr_fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def thin_border():
        s = Side(style="thin", color="FFE5E7EB")
        return Border(left=s, right=s, top=s, bottom=s)

    def header_font(color="FFFFFFFF", size=10):
        return Font(name="Calibri", bold=True, color=color, size=size)

    def body_font(color=GRAY_DARK, bold=False, size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    gender_str  = "Девочка" if child["gender"] == "girl" else "Мальчик"
    gender_icon = "♀" if child["gender"] == "girl" else "♂"
    age         = age_str(child["birthdate"])
    today_str   = date.today().strftime("%d.%m.%Y")

    # ══════════════════════════════════════════════════════════════════════
    # Лист 1: Общая информация
    # ══════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Общая информация"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 30

    # Шапка
    ws.merge_cells("A1:B1")
    ws["A1"] = "🌸 МамаБот — Медицинский отчёт"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    ws["A1"].fill = hdr_fill(PURPLE)
    ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:B2")
    ws["A2"] = f"Создан: {today_str}"
    ws["A2"].font = Font(name="Calibri", size=9, color="FFCCCCFF")
    ws["A2"].fill = hdr_fill(PURPLE)
    ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 8

    rows = [
        ("Имя ребёнка", f"{gender_icon} {child['name']}"),
        ("Дата рождения", child["birthdate"]),
        ("Возраст", age),
        ("Пол", gender_str),
    ]
    for i, (label, value) in enumerate(rows, start=4):
        ws.row_dimensions[i].height = 20
        ws[f"A{i}"] = label
        ws[f"A{i}"].font = body_font(color=GRAY_MID)
        ws[f"A{i}"].fill = hdr_fill(GRAY_LIGHT)
        ws[f"A{i}"].alignment = left()
        ws[f"A{i}"].border = thin_border()
        ws[f"B{i}"] = value
        ws[f"B{i}"].font = body_font(bold=True)
        ws[f"B{i}"].alignment = left()
        ws[f"B{i}"].border = thin_border()

    # ══════════════════════════════════════════════════════════════════════
    # Лист 2: Рост и вес
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Рост и вес")
    ws2.sheet_view.showGridLines = False
    for col, w in [("A", 16), ("B", 14), ("C", 14)]:
        ws2.column_dimensions[col].width = w

    ws2.merge_cells("A1:C1")
    ws2["A1"] = "📏 Рост и вес"
    ws2["A1"].font = header_font(size=12)
    ws2["A1"].fill = hdr_fill(PURPLE)
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 28

    headers = ["Дата", "Рост (см)", "Вес (кг)"]
    for ci, h in enumerate(headers, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = header_font()
        cell.fill = hdr_fill("FF818CF8")
        cell.alignment = center()
        cell.border = thin_border()
    ws2.row_dimensions[2].height = 20

    for ri, r in enumerate(growth_records or [], start=3):
        ws2.row_dimensions[ri].height = 18
        bg = PURPLE_LIGHT if ri % 2 == 1 else WHITE
        for ci, val in enumerate([r["date"], r["height_cm"] or "—", r["weight_kg"] or "—"], 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = body_font()
            cell.fill = hdr_fill(bg)
            cell.alignment = center()
            cell.border = thin_border()

    if not growth_records:
        ws2.merge_cells("A3:C3")
        ws2["A3"] = "Записей пока нет"
        ws2["A3"].font = body_font(color=GRAY_MID)
        ws2["A3"].alignment = center()

    # ══════════════════════════════════════════════════════════════════════
    # Лист 3: Прививки
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Прививки")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 48
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 14

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "💉 Прививки"
    ws3["A1"].font = header_font(size=12)
    ws3["A1"].fill = hdr_fill(GREEN)
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 28

    headers3 = ["Прививка", "Статус", "Дата / план"]
    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = header_font()
        cell.fill = hdr_fill("FF4ADE80")
        cell.alignment = center()
        cell.border = thin_border()
    ws3.row_dimensions[2].height = 20

    for ri, v in enumerate(vaccinations or [], start=3):
        ws3.row_dimensions[ri].height = 18
        if v["done_date"]:
            status, status_color, dv = "✓ Сделано", GREEN, v["done_date"]
            bg = GREEN_LIGHT
        else:
            try:
                d = datetime.strptime(v["scheduled_date"], "%d.%m.%Y").date()
                if d <= date.today():
                    status, status_color, bg = "⚠ Просрочено", RED, "FFFFF1F2"
                else:
                    status, status_color, bg = "→ Предстоит", PURPLE, PURPLE_LIGHT
            except Exception:
                status, status_color, bg = "→ Предстоит", PURPLE, PURPLE_LIGHT
            dv = v["scheduled_date"] or "—"

        bg = GREEN_LIGHT if ri % 2 == 1 else WHITE

        for ci, val in enumerate([v["vaccine_name"], status, dv], 1):
            cell = ws3.cell(row=ri, column=ci, value=val)
            cell.fill = hdr_fill(bg)
            cell.alignment = left() if ci == 1 else center()
            cell.border = thin_border()
            if ci == 2:
                cell.font = Font(name="Calibri", bold=True, color=status_color, size=10)
            else:
                cell.font = body_font()

    if not vaccinations:
        ws3.merge_cells("A3:C3")
        ws3["A3"] = "Прививки не добавлены"
        ws3["A3"].font = body_font(color=GRAY_MID)
        ws3["A3"].alignment = center()

    # ══════════════════════════════════════════════════════════════════════
    # Лист 4: Болезни
    # ══════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Болезни")
    ws4.sheet_view.showGridLines = False
    for col, w in [("A", 22), ("B", 14), ("C", 14), ("D", 14), ("E", 22), ("F", 22), ("G", 22)]:
        ws4.column_dimensions[col].width = w

    ws4.merge_cells("A1:G1")
    ws4["A1"] = "🤒 История болезней"
    ws4["A1"].font = header_font(size=12)
    ws4["A1"].fill = hdr_fill(AMBER)
    ws4["A1"].alignment = center()
    ws4.row_dimensions[1].height = 28

    headers4 = ["Болезнь", "Начало", "Конец", "Дата записи", "Температура", "Симптомы", "Лекарства"]
    for ci, h in enumerate(headers4, 1):
        cell = ws4.cell(row=2, column=ci, value=h)
        cell.font = header_font(color=GRAY_DARK)
        cell.fill = hdr_fill("FFFCD34D")
        cell.alignment = center()
        cell.border = thin_border()
    ws4.row_dimensions[2].height = 20

    row_idx = 3
    for ill in (illnesses or []):
        entries = db.get_illness_entries(ill["id"])
        if not entries:
            ws4.row_dimensions[row_idx].height = 18
            bg = AMBER_LIGHT if row_idx % 2 == 1 else WHITE
            for ci, val in enumerate([
                ill["illness_name"], ill["start_date"],
                ill["end_date"] or "активна", "—", "—", "—", "—"
            ], 1):
                cell = ws4.cell(row=row_idx, column=ci, value=val)
                cell.font = body_font()
                cell.fill = hdr_fill(bg)
                cell.alignment = left()
                cell.border = thin_border()
            row_idx += 1
        else:
            for e in entries:
                ws4.row_dimensions[row_idx].height = 18
                bg = AMBER_LIGHT if row_idx % 2 == 1 else WHITE
                hi = e["temperature"] and float(e["temperature"]) >= 38
                for ci, val in enumerate([
                    ill["illness_name"], ill["start_date"],
                    ill["end_date"] or "активна",
                    e["entry_date"],
                    f"{e['temperature']}°" if e["temperature"] else "—",
                    e["symptoms"] or "—",
                    e["medications_given"] or "—",
                ], 1):
                    cell = ws4.cell(row=row_idx, column=ci, value=val)
                    cell.fill = hdr_fill(bg)
                    cell.alignment = left()
                    cell.border = thin_border()
                    if ci == 5 and hi:
                        cell.font = Font(name="Calibri", bold=True, color=RED, size=10)
                    else:
                        cell.font = body_font()
                row_idx += 1

    if not illnesses:
        ws4.merge_cells("A3:G3")
        ws4["A3"] = "Болезней не записано"
        ws4["A3"].font = body_font(color=GRAY_MID)
        ws4["A3"].alignment = center()

    # ══════════════════════════════════════════════════════════════════════
    # Лист 5: Лекарства
    # ══════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Лекарства")
    ws5.sheet_view.showGridLines = False
    for col, w in [("A", 28), ("B", 16), ("C", 18), ("D", 16)]:
        ws5.column_dimensions[col].width = w

    ws5.merge_cells("A1:D1")
    ws5["A1"] = "💊 Лекарства"
    ws5["A1"].font = header_font(size=12)
    ws5["A1"].fill = hdr_fill(PINK)
    ws5["A1"].alignment = center()
    ws5.row_dimensions[1].height = 28

    headers5 = ["Название", "Доза", "Интервал", "До даты"]
    for ci, h in enumerate(headers5, 1):
        cell = ws5.cell(row=2, column=ci, value=h)
        cell.font = header_font()
        cell.fill = hdr_fill("FFF472B6")
        cell.alignment = center()
        cell.border = thin_border()
    ws5.row_dimensions[2].height = 20

    for ri, m in enumerate(medications or [], start=3):
        ws5.row_dimensions[ri].height = 18
        bg = PINK_LIGHT if ri % 2 == 1 else WHITE
        ih = int(m["interval_hours"])
        im = int((m["interval_hours"] - ih) * 60)
        ivl = f"каждые {ih} ч." if im == 0 else f"каждые {ih} ч. {im} мин."
        for ci, val in enumerate([m["name"], m["dose"] or "—", ivl, m["end_date"] or "—"], 1):
            cell = ws5.cell(row=ri, column=ci, value=val)
            cell.font = body_font()
            cell.fill = hdr_fill(bg)
            cell.alignment = left()
            cell.border = thin_border()

    if not medications:
        ws5.merge_cells("A3:D3")
        ws5["A3"] = "Лекарства не записаны"
        ws5["A3"].font = body_font(color=GRAY_MID)
        ws5["A3"].alignment = center()

    # ══════════════════════════════════════════════════════════════════════
    # Лист 6: Аллергии и противопоказания
    # ══════════════════════════════════════════════════════════════════════
    VIOLET      = "FF8B5CF6"
    VIOLET_LIGHT= "FFF5F3FF"

    ws6 = wb.create_sheet("Аллергии")
    ws6.sheet_view.showGridLines = False
    for col, w in [("A", 28), ("B", 30), ("C", 16)]:
        ws6.column_dimensions[col].width = w

    ws6.merge_cells("A1:C1")
    ws6["A1"] = "⚠️ Аллергии и противопоказания"
    ws6["A1"].font = header_font(size=12)
    ws6["A1"].fill = hdr_fill(VIOLET)
    ws6["A1"].alignment = center()
    ws6.row_dimensions[1].height = 28

    headers6 = ["Аллерген", "Реакция", "Тяжесть"]
    for ci, h in enumerate(headers6, 1):
        cell = ws6.cell(row=2, column=ci, value=h)
        cell.font = header_font(color=WHITE)
        cell.fill = hdr_fill("FFA78BFA")
        cell.alignment = center()
        cell.border = thin_border()
    ws6.row_dimensions[2].height = 20

    sev_map = {"mild": "🟡 Лёгкая", "moderate": "🟠 Средняя", "severe": "🔴 Тяжёлая"}
    sev_colors = {"mild": "FFF59E0B", "moderate": "FFF97316", "severe": "FFEF4444"}

    for ri, a in enumerate(allergies or [], start=3):
        ws6.row_dimensions[ri].height = 18
        bg = VIOLET_LIGHT if ri % 2 == 1 else WHITE
        sev_str = sev_map.get(a["severity"], "—")
        sev_col = sev_colors.get(a["severity"], GRAY_MID)
        for ci, val in enumerate([a["name"], a["reaction"] or "—", sev_str], 1):
            cell = ws6.cell(row=ri, column=ci, value=val)
            cell.fill = hdr_fill(bg)
            cell.alignment = left()
            cell.border = thin_border()
            if ci == 3:
                cell.font = Font(name="Calibri", bold=True, color=sev_col, size=10)
            else:
                cell.font = body_font()

    if not allergies:
        ws6.merge_cells("A3:C3")
        ws6["A3"] = "Аллергий не записано"
        ws6["A3"].font = body_font(color=GRAY_MID)
        ws6["A3"].alignment = center()

    # Противопоказания
    row_start = max(4, len(allergies or []) + 4)
    ws6.row_dimensions[row_start].height = 8
    row_start += 1

    ws6.merge_cells(f"A{row_start}:C{row_start}")
    ws6[f"A{row_start}"] = "🚫 Противопоказания"
    ws6[f"A{row_start}"].font = Font(name="Calibri", bold=True, color=VIOLET, size=11)
    ws6[f"A{row_start}"].alignment = left()
    ws6.row_dimensions[row_start].height = 22
    row_start += 1

    for ri, c in enumerate(contraindications or [], start=row_start):
        ws6.row_dimensions[ri].height = 18
        bg = VIOLET_LIGHT if ri % 2 == 1 else WHITE
        ws6.merge_cells(f"A{ri}:C{ri}")
        cell = ws6[f"A{ri}"]
        cell.value = f"• {c['name']}"
        cell.font = body_font()
        cell.fill = hdr_fill(bg)
        cell.alignment = left()
        cell.border = thin_border()

    if not contraindications:
        ws6.merge_cells(f"A{row_start}:C{row_start}")
        ws6[f"A{row_start}"] = "Противопоказаний не записано"
        ws6[f"A{row_start}"].font = body_font(color=GRAY_MID)
        ws6[f"A{row_start}"].alignment = center()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Telegram handlers ──────────────────────────────────────────────────────────

async def export_excel_select_child(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    user_id = update.effective_user.id
    children = db.get_children(user_id)
    is_cb = query is not None

    if is_cb:
        await query.answer()

    async def send(text, keyboard):
        markup = InlineKeyboardMarkup(keyboard)
        if is_cb:
            await query.edit_message_text(text, parse_mode="Markdown", reply_markup=markup)
        else:
            await update.message.reply_text(text, parse_mode="Markdown", reply_markup=markup)

    if not children:
        keyboard = [[InlineKeyboardButton("👶 Добавить ребёнка", callback_data="my_child")]]
        await send("📊 Нет детей для экспорта.", keyboard)
        return

    if len(children) == 1:
        if is_cb:
            await _do_excel_export(query, context, children[0]["id"], user_id)
        else:
            await _do_excel_export_msg(update.message, context, children[0]["id"], user_id)
        return

    keyboard = []
    for ch in children:
        emoji = "👧" if ch["gender"] == "girl" else "👦"
        keyboard.append([InlineKeyboardButton(
            f"{emoji} {ch['name']}", callback_data=f"excel_export:{ch['id']}"
        )])
    keyboard.append([InlineKeyboardButton("◀️ Назад", callback_data="my_child")])
    await send("📊 *Экспорт в Excel*\n\nВыберите ребёнка:", keyboard)


async def export_excel_for_child_cb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer("Создаю Excel...")
    child_id = int(query.data.split(":")[1])
    user_id = update.effective_user.id
    await _do_excel_export(query, context, child_id, user_id)


async def _do_excel_export_msg(message, context, child_id, user_id):
    ch = db.get_child(child_id, user_id)
    if not ch:
        await message.reply_text("Ребёнок не найден.")
        return
    wait = await message.reply_text("⏳ Создаю Excel-файл...")
    growth           = db.get_growth_records(child_id, user_id, limit=100)
    vaccines         = db.get_vaccinations(child_id, user_id)
    illnesses        = db.get_illnesses(child_id, active_only=False, limit=20)
    medications      = db.get_medications(child_id, active_only=False)
    allergies        = db.get_allergies(child_id)
    contraindications = db.get_contraindications(child_id)
    try:
        data = generate_child_xlsx(ch, growth, vaccines, illnesses, medications, allergies, contraindications)
        buf = io.BytesIO(data)
        buf.name = f"{ch['name']}_отчет.xlsx"
        await message.reply_document(document=buf, filename=f"{ch['name']}_отчет.xlsx",
                                     caption=f"📊 Excel-отчёт: {ch['name']}")
        await wait.delete()
    except Exception as e:
        await wait.edit_text(f"❌ Ошибка: {e}")


async def _do_excel_export(query, context, child_id, user_id):
    ch = db.get_child(child_id, user_id)
    if not ch:
        await query.edit_message_text("Ребёнок не найден.")
        return
    await query.edit_message_text("⏳ Создаю Excel-файл, подождите...")
    growth           = db.get_growth_records(child_id, user_id, limit=100)
    vaccines         = db.get_vaccinations(child_id, user_id)
    illnesses        = db.get_illnesses(child_id, active_only=False, limit=20)
    medications      = db.get_medications(child_id, active_only=False)
    allergies        = db.get_allergies(child_id)
    contraindications = db.get_contraindications(child_id)
    try:
        data = generate_child_xlsx(ch, growth, vaccines, illnesses, medications, allergies, contraindications)
        buf = io.BytesIO(data)
        buf.name = f"{ch['name']}_отчет.xlsx"
        keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data=f"child_view:{child_id}")]]
        await query.message.reply_document(
            document=buf, filename=f"{ch['name']}_отчет.xlsx",
            caption=f"📊 Excel-отчёт: {ch['name']}",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
        await query.delete_message()
    except Exception as e:
        keyboard = [[InlineKeyboardButton("◀️ Назад", callback_data=f"child_view:{child_id}")]]
        await query.edit_message_text(f"❌ Ошибка создания Excel.\n\n`{e}`",
                                      parse_mode="Markdown",
                                      reply_markup=InlineKeyboardMarkup(keyboard))
